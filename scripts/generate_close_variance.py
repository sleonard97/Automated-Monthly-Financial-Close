"""Generate a monthly close variance review workbook.

This v1 reads:
- an EdTec GL Detail Report export
- a school financial model workbook

It writes an Excel workbook with Summary, Detail, and Problems tabs.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import warnings
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - dependency guard
    raise SystemExit(
        "Missing dependency: openpyxl. Run `python -m pip install -r requirements.txt`."
    ) from exc


XML_NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}

GL_REQUIRED_HEADERS = {
    "account",
    "type",
    "date",
    "debit",
    "credit",
    "balance",
    "description",
    "memo",
    "account number",
}

FISCAL_MONTH_INDEX = {
    "jul": 1,
    "july": 1,
    "aug": 2,
    "august": 2,
    "sep": 3,
    "sept": 3,
    "september": 3,
    "oct": 4,
    "october": 4,
    "nov": 5,
    "november": 5,
    "dec": 6,
    "december": 6,
    "jan": 7,
    "january": 7,
    "feb": 8,
    "february": 8,
    "mar": 9,
    "march": 9,
    "apr": 10,
    "april": 10,
    "may": 11,
    "jun": 12,
    "june": 12,
}

MONTH_DISPLAY = {
    "jul": "July",
    "july": "July",
    "aug": "August",
    "august": "August",
    "sep": "September",
    "sept": "September",
    "september": "September",
    "oct": "October",
    "october": "October",
    "nov": "November",
    "november": "November",
    "dec": "December",
    "december": "December",
    "jan": "January",
    "january": "January",
    "feb": "February",
    "february": "February",
    "mar": "March",
    "march": "March",
    "apr": "April",
    "april": "April",
    "may": "May",
    "jun": "June",
    "june": "June",
}


@dataclass
class AccountActual:
    account_number: str
    account_name: str = ""
    raw_net: float = 0.0
    debit_total: float = 0.0
    credit_total: float = 0.0
    transaction_count: int = 0
    latest_activity_date: dt.datetime | dt.date | None = None
    rolled_from: list[str] = field(default_factory=list)


@dataclass
class ForecastAccount:
    account_number: str
    account_name: str
    current_forecast: float
    model_section: str
    account_type: str
    source_row: int
    actual_ytd: float = 0.0


@dataclass
class GLParseResult:
    headers: list[str]
    detail_rows: list[dict[str, Any]]
    actuals: dict[str, AccountActual]
    source_title: str = ""
    source_period: str = ""
    problems: list[dict[str, str]] = field(default_factory=list)


@dataclass
class ModelParseResult:
    school_name: str
    close_text: str
    close_month: str
    usage_benchmark: float
    forecasts: dict[str, ForecastAccount]
    problems: list[dict[str, str]] = field(default_factory=list)


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def clean_path(value: str) -> Path:
    return Path(value.strip().strip('"').strip("'")).expanduser()


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("$", "").replace(",", "")
    if text in {"-", "--"}:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative else number


def normalize_account_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        if abs(float(value) - int(float(value))) < 0.000001:
            return str(int(float(value)))
        return str(value).strip()

    text = str(value).strip()
    if not text:
        return ""
    text = text.replace(",", "")
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    if re.fullmatch(r"\d+", text):
        return text

    match = re.match(r"^\s*(\d{3,8})\s*(?:-|$)", text)
    if match:
        return match.group(1)
    return ""


def split_account_label(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    match = re.match(r"^\s*(\d{3,8})\s*-\s*(.+)$", text)
    if not match:
        return "", text
    return match.group(1), match.group(2).strip()


def xml_cell_value(cell: ET.Element) -> Any:
    data = cell.find("ss:Data", XML_NS)
    if data is None:
        return None
    value = data.text or ""
    data_type = data.attrib.get(
        "{urn:schemas-microsoft-com:office:spreadsheet}Type", ""
    )
    if data_type == "Number":
        return to_number(value)
    if data_type == "DateTime":
        try:
            return dt.datetime.fromisoformat(value.replace("Z", "+00:00")).replace(
                tzinfo=None
            )
        except ValueError:
            return value
    return value


def iter_excel_xml_rows(path: Path) -> list[tuple[int, list[Any]]]:
    root = ET.parse(path).getroot()
    worksheet = root.find("ss:Worksheet", XML_NS)
    if worksheet is None:
        raise ValueError("No worksheet found in XML-style Excel file.")
    table = worksheet.find("ss:Table", XML_NS)
    if table is None:
        raise ValueError("No table found in XML-style Excel worksheet.")

    parsed_rows: list[tuple[int, list[Any]]] = []
    physical_row = 0
    for row in table.findall("ss:Row", XML_NS):
        row_index = row.attrib.get(
            "{urn:schemas-microsoft-com:office:spreadsheet}Index"
        )
        physical_row = int(row_index) if row_index and row_index.isdigit() else physical_row + 1

        values: list[Any] = []
        current_col = 0
        for cell in row.findall("ss:Cell", XML_NS):
            col_index = cell.attrib.get(
                "{urn:schemas-microsoft-com:office:spreadsheet}Index"
            )
            if col_index and col_index.isdigit():
                while current_col < int(col_index) - 1:
                    values.append(None)
                    current_col += 1
            values.append(xml_cell_value(cell))
            current_col += 1
        parsed_rows.append((physical_row, values))
    return parsed_rows


def read_xlsx_like_rows(path: Path) -> list[tuple[int, list[Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet = workbook.worksheets[0]
        rows: list[tuple[int, list[Any]]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append((len(rows) + 1, list(row)))
        return rows
    finally:
        workbook.close()


def load_gl_detail(path: Path) -> GLParseResult:
    if not path.exists():
        raise FileNotFoundError(f"GL detail file not found: {path}")

    first_bytes = path.read_bytes()[:32]
    if first_bytes.lstrip().startswith(b"<?xml"):
        source_rows = iter_excel_xml_rows(path)
    elif path.suffix.lower() in {".xlsx", ".xlsm"}:
        source_rows = read_xlsx_like_rows(path)
    else:
        raise ValueError(
            "Unsupported GL detail format. Use the XML-style EdTec .xls export or .xlsx."
        )

    title_rows = [row for _, row in source_rows[:5]]
    source_title = str(title_rows[2][0]) if len(title_rows) > 2 and title_rows[2] else ""
    source_period = str(title_rows[3][0]) if len(title_rows) > 3 and title_rows[3] else ""

    header_index = -1
    headers: list[str] = []
    for index, (_, values) in enumerate(source_rows):
        normalized = {normalize_header(value) for value in values}
        if GL_REQUIRED_HEADERS.issubset(normalized):
            header_index = index
            headers = [str(value or "").strip() for value in values]
            break
    if header_index < 0:
        raise ValueError("Could not find the GL detail header row.")

    actuals: dict[str, AccountActual] = {}
    detail_rows: list[dict[str, Any]] = []
    problems: list[dict[str, str]] = []
    current_account_number = ""
    current_account_name = ""

    for source_row_number, values in source_rows[header_index + 1 :]:
        padded = values + [None] * max(0, len(headers) - len(values))
        row_dict = {
            headers[col_index]: padded[col_index] if col_index < len(padded) else None
            for col_index in range(len(headers))
        }

        if not any(value not in (None, "") for value in row_dict.values()):
            continue

        account_label = str(row_dict.get("Account") or "").strip()
        label_number, label_name = split_account_label(account_label)
        if label_number and not str(row_dict.get("Type") or "").strip():
            current_account_number = label_number
            current_account_name = label_name

        row_dict["Source Row"] = source_row_number
        row_dict["Current Account Number"] = current_account_number
        row_dict["Current Account Name"] = current_account_name
        detail_rows.append(row_dict)

        account_number = normalize_account_number(row_dict.get("Account Number"))
        if not account_number:
            continue

        debit = to_number(row_dict.get("Debit"))
        credit = to_number(row_dict.get("Credit"))
        if debit == 0 and credit == 0:
            continue

        account = actuals.setdefault(
            account_number,
            AccountActual(
                account_number=account_number,
                account_name=current_account_name,
            ),
        )
        if not account.account_name:
            account.account_name = current_account_name
        account.raw_net += debit - credit
        account.debit_total += debit
        account.credit_total += credit
        account.transaction_count += 1

        date_value = row_dict.get("Date")
        if isinstance(date_value, (dt.datetime, dt.date)):
            if account.latest_activity_date is None or date_value > account.latest_activity_date:
                account.latest_activity_date = date_value

    if not actuals:
        problems.append(
            {
                "Severity": "Error",
                "Type": "Import",
                "Account Number": "",
                "Account Name": "",
                "Message": "No transaction rows with account numbers and debit/credit values were found.",
            }
        )

    return GLParseResult(
        headers=headers,
        detail_rows=detail_rows,
        actuals=actuals,
        source_title=source_title,
        source_period=source_period,
        problems=problems,
    )


def detect_close_month(close_text: str) -> tuple[str, float]:
    match = re.search(r"\bAs of\s+([A-Za-z]+)\b", close_text or "", re.IGNORECASE)
    if not match:
        match = re.search(
            r"\b(January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\b",
            close_text or "",
            re.IGNORECASE,
        )
    if not match:
        return "", 0.0
    key = match.group(1).lower()
    fiscal_index = FISCAL_MONTH_INDEX.get(key, 0)
    if not fiscal_index:
        return "", 0.0
    return MONTH_DISPLAY[key], fiscal_index / 12.0


def classify_section(label: str, current_type: str) -> str:
    text = label.upper()
    if "REVENUE" in text or "LCFF" in text or "FUNDRAISING" in text:
        return "Revenue"
    if "EXPENSE" in text or "SALAR" in text or "BENEFIT" in text:
        return "Expense"
    if "CAPITAL" in text or "CAPEX" in text or "FIXED ASSET" in text:
        return "CapEx"
    return current_type or "Other"


def looks_like_section_label(col_a: Any, col_b: Any, col_c: Any) -> bool:
    if col_a in (None, ""):
        return False
    if normalize_account_number(col_a):
        return False
    text = str(col_a).strip()
    if not text or text.upper().startswith("SUBTOTAL"):
        return False
    if col_b not in (None, ""):
        return False
    if to_number(col_c) != 0:
        return False
    return True


def find_header_column(sheet, header_name: str, header_row: int = 6) -> int:
    target = normalize_header(header_name)
    for col_index in range(1, sheet.max_column + 1):
        if normalize_header(sheet.cell(header_row, col_index).value) == target:
            return col_index
    raise ValueError(f"Could not find `{header_name}` on the YTD tab.")


def has_visible_ytd_data(*values: Any) -> bool:
    return any(abs(to_number(value)) > 0.000001 for value in values)


def load_forecast_model(path: Path) -> ModelParseResult:
    if not path.exists():
        raise FileNotFoundError(f"Financial model file not found: {path}")

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Conditional Formatting extension.*")
        warnings.filterwarnings("ignore", message="Data Validation extension.*")
        workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    try:
        if "YTD" not in workbook.sheetnames:
            raise ValueError("Financial model is missing a `YTD` tab.")
        ytd = workbook["YTD"]
        school_name = str(ytd["A1"].value or "").strip()

        close_text = ""
        for sheet_name in ("YTD", "Detail"):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                for value in row:
                    text = str(value or "")
                    if "As of" in text:
                        close_text = text.strip()
                        break
                if close_text:
                    break
            if close_text:
                break

        close_month, usage_benchmark = detect_close_month(close_text)
        problems: list[dict[str, str]] = []
        if not usage_benchmark:
            problems.append(
                {
                    "Severity": "Warning",
                    "Type": "Close Month",
                    "Account Number": "",
                    "Account Name": "",
                    "Message": "Could not detect close month from the model. Usage benchmark is blank.",
                }
            )

        forecasts: dict[str, ForecastAccount] = {}
        current_section = ""
        current_type = "Other"
        actual_ytd_col = find_header_column(ytd, "Actual YTD")
        current_forecast_col = find_header_column(ytd, "Current Forecast")
        hidden_account_count = 0

        for row_number in range(1, ytd.max_row + 1):
            col_a = ytd.cell(row_number, 1).value
            col_b = ytd.cell(row_number, 2).value
            actual_ytd = to_number(ytd.cell(row_number, actual_ytd_col).value)
            current_forecast = to_number(ytd.cell(row_number, current_forecast_col).value)

            if looks_like_section_label(col_a, col_b, current_forecast):
                current_section = str(col_a).strip()
                current_type = classify_section(current_section, current_type)
                continue

            account_number = normalize_account_number(col_a)
            account_name = str(col_b or "").strip()
            if not account_number or not account_name:
                continue
            if account_name.upper().startswith("SUBTOTAL"):
                continue
            if ytd.row_dimensions[row_number].hidden:
                hidden_account_count += 1
                continue
            if not has_visible_ytd_data(actual_ytd, current_forecast):
                continue

            existing = forecasts.get(account_number)
            if existing:
                existing.current_forecast += current_forecast
                existing.actual_ytd += actual_ytd
                if current_section and current_section not in existing.model_section:
                    existing.model_section = f"{existing.model_section}; {current_section}"
            else:
                forecasts[account_number] = ForecastAccount(
                    account_number=account_number,
                    account_name=account_name,
                    current_forecast=current_forecast,
                    model_section=current_section,
                    account_type=current_type,
                    source_row=row_number,
                    actual_ytd=actual_ytd,
                )

        if not forecasts:
            problems.append(
                {
                    "Severity": "Error",
                    "Type": "Forecast",
                    "Account Number": "",
                    "Account Name": "",
                    "Message": "No visible account-level rows with YTD data were found on the YTD tab.",
                }
            )
        if hidden_account_count:
            problems.append(
                {
                    "Severity": "Info",
                    "Type": "Forecast Scope",
                    "Account Number": "",
                    "Account Name": "",
                    "Message": f"Excluded {hidden_account_count:,} hidden YTD account rows from the forecast universe.",
                }
            )

        return ModelParseResult(
            school_name=school_name,
            close_text=close_text,
            close_month=close_month,
            usage_benchmark=usage_benchmark,
            forecasts=forecasts,
            problems=problems,
        )
    finally:
        workbook.close()


def comparable_actual(raw_net: float, account_type: str) -> float:
    if account_type == "Revenue":
        return -raw_net
    return raw_net


def account_type_from_account_number(account_number: str) -> str:
    if account_number.startswith("8"):
        return "Revenue"
    if account_number[:1] in {"1", "2", "3", "4", "5", "6", "7"}:
        return "Expense"
    return "Other"


def is_balance_sheet_account(account_number: str) -> bool:
    return account_number.startswith("9")


def is_revenue_or_expense_account(account_number: str) -> bool:
    return account_type_from_account_number(account_number) in {"Revenue", "Expense"}


def resolve_benefit_rollup_account(
    account_number: str, forecasts: dict[str, ForecastAccount]
) -> str:
    """Roll detailed benefits GL accounts to the visible forecast parent account."""
    if account_number in forecasts:
        return account_number
    if not re.fullmatch(r"3\d{3}", account_number):
        return account_number

    parent_account = f"{account_number[:2]}00"
    parent_forecast = forecasts.get(parent_account)
    if not parent_forecast:
        return account_number

    parent_context = f"{parent_forecast.model_section} {parent_forecast.account_name}".lower()
    benefit_terms = (
        "benefit",
        "medicare",
        "oasdi",
        "insurance",
        "unemployment",
        "workers comp",
    )
    if any(term in parent_context for term in benefit_terms):
        return parent_account
    return account_number


def combine_account_actual(target: AccountActual, source: AccountActual) -> None:
    target.raw_net += source.raw_net
    target.debit_total += source.debit_total
    target.credit_total += source.credit_total
    target.transaction_count += source.transaction_count
    if not target.account_name:
        target.account_name = source.account_name
    if (
        source.latest_activity_date is not None
        and (
            target.latest_activity_date is None
            or source.latest_activity_date > target.latest_activity_date
        )
    ):
        target.latest_activity_date = source.latest_activity_date


def rollup_actuals_to_forecast_accounts(
    actuals: dict[str, AccountActual], forecasts: dict[str, ForecastAccount]
) -> tuple[dict[str, AccountActual], list[dict[str, str]]]:
    rolled_actuals: dict[str, AccountActual] = {}
    rollup_notes: list[dict[str, str]] = []

    for source_account_number, source in actuals.items():
        target_account_number = resolve_benefit_rollup_account(
            source_account_number, forecasts
        )
        target = rolled_actuals.setdefault(
            target_account_number,
            AccountActual(
                account_number=target_account_number,
                account_name=source.account_name,
            ),
        )
        combine_account_actual(target, source)

        if target_account_number != source_account_number:
            source_label = f"{source_account_number} - {source.account_name}".strip(" -")
            target.rolled_from.append(source_label)
            target_forecast = forecasts[target_account_number]
            rollup_notes.append(
                {
                    "Severity": "Info",
                    "Type": "Benefit Rollup",
                    "Account Number": target_account_number,
                    "Account Name": target_forecast.account_name,
                    "Message": (
                        f"Rolled GL account {source_label} into visible YTD forecast "
                        f"account {target_account_number} - {target_forecast.account_name}."
                    ),
                }
            )

    return rolled_actuals, rollup_notes


def percent(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.1%}"


def money(value: float | None) -> str:
    if value is None:
        return ""
    return f"${value:,.0f}"


def make_actual_only_review_prompt(account_type: str) -> str:
    return (
        f"Review because this {account_type.lower()} account has YTD activity in the GL file "
        "but is not listed on a visible YTD forecast row. Confirm whether it should be added, "
        "mapped, or intentionally excluded."
    )


def make_review_prompt(
    account_type: str,
    pct_used: float | None,
    close_month: str,
    benchmark: float | None,
    current_forecast: float,
) -> str:
    if current_forecast == 0:
        return (
            "Review because this account has YTD activity but no current forecast. "
            "Confirm mapping and forecast assumptions."
        )
    if pct_used is None or benchmark is None:
        return (
            "Review because this account was flagged by the variance rules. "
            "Confirm whether the forecast and actuals are mapped correctly."
        )
    month_text = close_month or "the close month"
    return (
        f"Review because this {account_type.lower()} account has used {percent(pct_used)} "
        f"of the full-year forecast by {month_text}, compared with the "
        f"{percent(benchmark)} month-based benchmark. Confirm whether remaining forecast is sufficient."
    )


def build_review_rows(
    gl: GLParseResult,
    model: ModelParseResult,
    materiality_dollars: float,
    materiality_percent: float,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    review_rows: list[dict[str, Any]] = []
    problems: list[dict[str, str]] = []
    rolled_actuals, rollup_notes = rollup_actuals_to_forecast_accounts(
        gl.actuals, model.forecasts
    )
    problems.extend(rollup_notes)

    actual_accounts = set(rolled_actuals)
    forecast_accounts = set(model.forecasts)

    for account_number in sorted(actual_accounts - forecast_accounts):
        if is_balance_sheet_account(account_number):
            continue
        if not is_revenue_or_expense_account(account_number):
            continue
        actual = rolled_actuals[account_number]
        account_type = account_type_from_account_number(account_number)
        actual_ytd = comparable_actual(actual.raw_net, account_type)
        severity = "Warning" if abs(actual.raw_net) >= materiality_dollars else "Info"
        problems.append(
            {
                "Severity": severity,
                "Type": "Unmatched Actual",
                "Account Number": account_number,
                "Account Name": actual.account_name,
                "Message": "Account appears in GL actuals but was not found on a visible YTD forecast row.",
            }
        )
        review_rows.append(
            {
                "Account Number": account_number,
                "Account Name": actual.account_name,
                "Account Type": account_type,
                "Model Section": "Actual-only GL account",
                "_Forecast Row": 999999,
                "YTD Actual": actual_ytd,
                "Current Forecast": 0,
                "Remaining $": -actual_ytd,
                "% Forecast Used": None,
                "Month Benchmark": model.usage_benchmark or None,
                "Benchmark Variance $": None,
                "Benchmark Variance %": None,
                "Favorable / Unfavorable": "Review",
                "Flag Reason": "Actual activity exists in GL but account is not listed on a visible YTD forecast row.",
                "Rolled Actual Accounts": "; ".join(actual.rolled_from),
                "AI Review Prompt": make_actual_only_review_prompt(account_type),
                "Transaction Count": actual.transaction_count,
                "Latest Activity Date": actual.latest_activity_date,
            }
        )

    for account_number in sorted(actual_accounts & forecast_accounts):
        actual = rolled_actuals[account_number]
        forecast = model.forecasts[account_number]
        account_type = forecast.account_type
        actual_ytd = comparable_actual(actual.raw_net, account_type)
        current_forecast = abs(forecast.current_forecast)
        remaining = current_forecast - actual_ytd
        pct_used = actual_ytd / current_forecast if current_forecast else None
        benchmark = model.usage_benchmark or None
        expected_ytd = current_forecast * benchmark if benchmark is not None else None
        benchmark_variance = actual_ytd - expected_ytd if expected_ytd is not None else None
        benchmark_variance_pct = (
            pct_used - benchmark if pct_used is not None and benchmark is not None else None
        )

        flag_reason = ""
        favorability = "Review"

        if current_forecast == 0 and abs(actual_ytd) >= materiality_dollars:
            flag_reason = "Actual activity exists with no current forecast."
        elif benchmark_variance is not None and benchmark_variance_pct is not None:
            material = (
                abs(benchmark_variance) >= materiality_dollars
                and abs(benchmark_variance_pct) >= materiality_percent
            )
            if material:
                direction = "above" if benchmark_variance_pct > 0 else "below"
                flag_reason = (
                    f"Percent used is {percent(abs(benchmark_variance_pct))} {direction} "
                    "the month-based benchmark."
                )
                if account_type == "Revenue":
                    favorability = "Unfavorable" if benchmark_variance_pct < 0 else "Favorable"
                elif account_type in {"Expense", "CapEx"}:
                    favorability = "Unfavorable" if benchmark_variance_pct > 0 else "Favorable"

        if not flag_reason:
            continue

        review_rows.append(
            {
                "Account Number": account_number,
                "Account Name": forecast.account_name or actual.account_name,
                "Account Type": account_type,
                "Model Section": forecast.model_section,
                "_Forecast Row": forecast.source_row,
                "YTD Actual": actual_ytd,
                "Current Forecast": current_forecast,
                "Remaining $": remaining,
                "% Forecast Used": pct_used,
                "Month Benchmark": benchmark,
                "Benchmark Variance $": benchmark_variance,
                "Benchmark Variance %": benchmark_variance_pct,
                "Favorable / Unfavorable": favorability,
                "Flag Reason": flag_reason,
                "Rolled Actual Accounts": "; ".join(actual.rolled_from),
                "AI Review Prompt": make_review_prompt(
                    account_type=account_type,
                    pct_used=pct_used,
                    close_month=model.close_month,
                    benchmark=benchmark,
                    current_forecast=current_forecast,
                ),
                "Transaction Count": actual.transaction_count,
                "Latest Activity Date": actual.latest_activity_date,
            }
        )

    review_rows.sort(
        key=lambda row: (
            int(row.get("_Forecast Row") or 999999),
            str(row.get("Account Number") or ""),
        ),
    )
    return review_rows, problems


def add_table(sheet, start_row: int, start_col: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row:
        return
    ref = f"{sheet.cell(start_row, start_col).coordinate}:{sheet.cell(end_row, end_col).coordinate}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def set_common_widths(sheet, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_output_workbook(
    output_path: Path,
    gl: GLParseResult,
    model: ModelParseResult,
    review_rows: list[dict[str, Any]],
    problems: list[dict[str, str]],
    materiality_dollars: float,
    materiality_percent: float,
    gl_path: Path,
    model_path: Path,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    detail = workbook.create_sheet("Detail")
    problems_sheet = workbook.create_sheet("Problems")

    write_summary_sheet(
        summary,
        review_rows,
        model,
        gl_path,
        model_path,
        materiality_dollars,
        materiality_percent,
    )
    write_detail_sheet(detail, gl)
    write_problems_sheet(problems_sheet, gl, model, problems)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_summary_sheet(
    sheet,
    review_rows: list[dict[str, Any]],
    model: ModelParseResult,
    gl_path: Path,
    model_path: Path,
    materiality_dollars: float,
    materiality_percent: float,
) -> None:
    headers = [
        "Account Number",
        "Account Name",
        "YTD Actual",
        "Current Forecast",
        "Remaining $",
        "% Forecast Used",
        "Benchmark Variance $",
        "Benchmark Variance %",
        "Transaction Count",
        "Latest Activity Date",
    ]
    sheet["A1"] = "Monthly Close Variance Review"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A2"] = "School"
    sheet["B2"] = model.school_name
    sheet["A3"] = "Close Month"
    sheet["B3"] = model.close_month or model.close_text
    sheet["A4"] = "Usage Benchmark"
    sheet["B4"] = model.usage_benchmark
    sheet["A5"] = "Materiality"
    sheet["B5"] = f"{money(materiality_dollars)} and {percent(materiality_percent)}"
    sheet["A6"] = "Source Files"
    sheet["B6"] = f"{gl_path.name}; {model_path.name}"
    sheet["B4"].number_format = "0.0%"

    revenue_rows = [
        row for row in review_rows if account_type_from_account_number(str(row.get("Account Number") or "")) == "Revenue"
    ]
    expense_rows = [
        row for row in review_rows if account_type_from_account_number(str(row.get("Account Number") or "")) == "Expense"
    ]

    first_section_row = 8
    next_row = write_summary_section(
        sheet=sheet,
        title="Revenue Accounts",
        start_row=first_section_row,
        headers=headers,
        rows=revenue_rows,
        table_name="SummaryRevenue",
        empty_message="No revenue accounts met the v1 flagging rules.",
        section_type="Revenue",
    )
    write_summary_section(
        sheet=sheet,
        title="Expense Accounts",
        start_row=next_row + 2,
        headers=headers,
        rows=expense_rows,
        table_name="SummaryExpense",
        empty_message="No expense accounts met the v1 flagging rules.",
        section_type="Expense",
    )

    sheet.freeze_panes = f"A{first_section_row}"
    set_common_widths(
        sheet,
        {
            "A": 16,
            "B": 38,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 18,
            "H": 16,
            "I": 16,
            "J": 18,
        },
    )


def write_summary_section(
    sheet,
    title: str,
    start_row: int,
    headers: list[str],
    rows: list[dict[str, Any]],
    table_name: str,
    empty_message: str,
    section_type: str,
) -> int:
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.cell(start_row, 1).value = title
    sheet.cell(start_row, 1).font = Font(bold=True, size=12)
    sheet.cell(start_row, 1).fill = title_fill
    sheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=len(headers),
    )

    header_row = start_row + 1
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(header_row, col_index).value = header
    style_header(sheet[header_row])

    if rows:
        for row_index, row in enumerate(rows, start=header_row + 1):
            for col_index, header in enumerate(headers, start=1):
                sheet.cell(row_index, col_index).value = row.get(header)
        last_row = header_row + len(rows)
        add_table(sheet, header_row, 1, last_row, len(headers), table_name)
        format_summary_data_rows(sheet, header_row + 1, last_row, section_type)
        return last_row + 1

    message_row = header_row + 1
    sheet.cell(message_row, 1).value = empty_message
    sheet.cell(message_row, 1).font = Font(italic=True, color="666666")
    return message_row + 1


def format_summary_data_rows(sheet, first_row: int, last_row: int, section_type: str) -> None:
    currency_cols = ["C", "D", "E"]
    for row in range(first_row, last_row + 1):
        for col in currency_cols:
            sheet[f"{col}{row}"].number_format = '$#,##0;[Red]($#,##0);-'
        sheet[f"F{row}"].number_format = "0.0%;[Red](0.0%);-"
        if section_type == "Expense":
            sheet[f"G{row}"].number_format = '[Red]$#,##0;($#,##0);-'
            sheet[f"H{row}"].number_format = "[Red]0.0%;(0.0%);-"
            color = "C00000" if to_number(sheet[f"G{row}"].value) > 0 else "000000"
            for col in ("A", "B"):
                sheet[f"{col}{row}"].font = Font(color=color)
        else:
            sheet[f"G{row}"].number_format = '$#,##0;[Red]($#,##0);-'
            sheet[f"H{row}"].number_format = "0.0%;[Red](0.0%);-"
        sheet[f"J{row}"].number_format = "yyyy-mm-dd"


def write_detail_sheet(sheet, gl: GLParseResult) -> None:
    helper_headers = ["Source Row", "Current Account Number", "Current Account Name"]
    headers = helper_headers + [header for header in gl.headers if header not in helper_headers]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(1, col_index).value = header
    style_header(sheet[1])

    for row_index, row in enumerate(gl.detail_rows, start=2):
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row_index, col_index).value = row.get(header)

    if gl.detail_rows:
        add_table(sheet, 1, 1, len(gl.detail_rows) + 1, len(headers), "GLDetail")

    sheet.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        letter = sheet.cell(1, col).column_letter
        sheet.column_dimensions[letter].width = 18


def build_tie_out_rows(gl: GLParseResult) -> list[dict[str, str]]:
    detail_net = 0.0
    for row in gl.detail_rows:
        account_number = normalize_account_number(row.get("Account Number"))
        if not account_number:
            continue
        detail_net += to_number(row.get("Debit")) - to_number(row.get("Credit"))

    aggregate_net = sum(actual.raw_net for actual in gl.actuals.values())
    difference = detail_net - aggregate_net
    severity = "OK" if abs(difference) < 0.01 else "Warning"
    return [
        {
            "Severity": severity,
            "Type": "Total Actuals Tie-Out",
            "Account Number": "",
            "Account Name": "",
            "Message": (
                f"Imported GL detail net activity ({money(detail_net)}) ties to account aggregates "
                f"({money(aggregate_net)}) with difference {money(difference)}."
            ),
        }
    ]


def write_problems_sheet(
    sheet,
    gl: GLParseResult,
    model: ModelParseResult,
    problems: list[dict[str, str]],
) -> None:
    headers = ["Severity", "Type", "Account Number", "Account Name", "Message"]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(1, col_index).value = header
    style_header(sheet[1])

    all_problems: list[dict[str, str]] = []
    all_problems.extend(gl.problems)
    all_problems.extend(model.problems)
    all_problems.extend(build_tie_out_rows(gl))
    all_problems.extend(problems)

    for row_index, row in enumerate(all_problems, start=2):
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row_index, col_index).value = row.get(header, "")
        sheet.cell(row_index, 5).alignment = Alignment(wrap_text=True, vertical="top")

    if not all_problems:
        sheet.cell(2, 1).value = "OK"
        sheet.cell(2, 2).value = "Validation"
        sheet.cell(2, 5).value = "No issues were identified by the v1 checks."

    add_table(sheet, 1, 1, max(2, len(all_problems) + 1), len(headers), "Problems")
    sheet.freeze_panes = "A2"
    set_common_widths(
        sheet,
        {"A": 12, "B": 22, "C": 18, "D": 38, "E": 90},
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a monthly close variance analysis workbook."
    )
    parser.add_argument("--gl", help="Path to the EdTec GL detail export.")
    parser.add_argument("--model", help="Path to the financial model workbook.")
    parser.add_argument("--output", help="Path for the generated .xlsx workbook.")
    parser.add_argument(
        "--materiality-dollars",
        type=float,
        default=5000.0,
        help="Dollar materiality threshold. Default: 5000.",
    )
    parser.add_argument(
        "--materiality-percent",
        type=float,
        default=0.10,
        help="Percent materiality threshold as a decimal. Default: 0.10.",
    )
    return parser.parse_args(argv)


def prompt_for_missing_paths(args: argparse.Namespace) -> tuple[Path, Path, Path]:
    gl_path = clean_path(args.gl or input("Path to GL detail export: "))
    model_path = clean_path(args.model or input("Path to financial model workbook: "))
    output_path = clean_path(args.output or input("Save output workbook as (.xlsx): "))
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    return gl_path, model_path, output_path


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    gl_path, model_path, output_path = prompt_for_missing_paths(args)

    print("Reading GL detail export...")
    gl = load_gl_detail(gl_path)
    print(f"  Found {len(gl.actuals):,} actual accounts.")

    print("Reading financial model...")
    model = load_forecast_model(model_path)
    print(f"  Found {len(model.forecasts):,} forecast accounts.")

    print("Building variance review...")
    review_rows, problems = build_review_rows(
        gl,
        model,
        materiality_dollars=args.materiality_dollars,
        materiality_percent=args.materiality_percent,
    )
    print(f"  Flagged {len(review_rows):,} accounts for review.")
    print(f"  Logged {len(problems):,} match/problem rows.")

    print("Writing output workbook...")
    write_output_workbook(
        output_path=output_path,
        gl=gl,
        model=model,
        review_rows=review_rows,
        problems=problems,
        materiality_dollars=args.materiality_dollars,
        materiality_percent=args.materiality_percent,
        gl_path=gl_path,
        model_path=model_path,
    )
    print(f"Done: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
